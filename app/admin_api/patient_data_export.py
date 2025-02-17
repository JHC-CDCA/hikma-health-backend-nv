from admin_api.patient_data_import import PatientDataRow, COLUMNS
from visits.data_access import all_visits
from openpyxl import load_workbook
from events.data_access import events_by_visit, patient_details
from patients.data_access import patient_from_id
from events.event_export import (write_vitals_event, write_medical_hx_event, write_evaluation_event, write_patient_details_event,
                                 write_med_stock_event, write_med_otc_event, write_controlled_med_event, write_med_pathologies_event, write_psych_pathologies_event, write_household_event,
                                 write_lab_orders_event, write_lab_tests_event, write_urine_tests_event, write_pap_event, write_ultrasound_event, write_family_planning_event,
                                 write_dental_origin_event, write_dental_treatment_event, write_program_trainings_event, write_xray_orders_event, write_xray_results_event,
                                 write_optometry_event, write_accident_report_event, write_nursing_care_event)
from datetime import datetime, timedelta, date
from tempfile import NamedTemporaryFile
import json
from google.cloud import storage
from config import EXPORTS_STORAGE_BUCKET


def most_recent_export():
    storage_client = storage.Client()
    blobs = storage_client.list_blobs(EXPORTS_STORAGE_BUCKET)
    most_recent = max(blobs, key=lambda b: b.name)
    output = NamedTemporaryFile('wb', suffix='.xlsx', delete=False)
    output.close()
    most_recent.download_to_filename(output.name)
    return output.name


class PatientDataExporter:
    def __init__(self):
        self.rows = []

    def run(self):
        workbook = load_workbook('data/base_export.xlsx')
        worksheet = workbook.get_sheet_by_name('Sheet1')
        for i, row in enumerate(self.iter_data_rows()):
            self.write_row(worksheet, i, row)
        output = NamedTemporaryFile('wb', suffix='.xlsx', delete=False)
        output.close()
        workbook.save(output.name)
        return output.name

    def write_row(self, worksheet, row_index, row):
        for col_index, col_key in enumerate(COLUMNS):
            value = getattr(row, col_key)
            if value is not None:
                cell = worksheet.cell(row_index + 3, col_index + 1)
                cell.value = value

    def iter_data_rows(self):
        for visit in all_visits():
            if not visit.patient_id:
                continue
            patient = patient_from_id(visit.patient_id)
            if not patient:
                continue
            row = PatientDataRow(
                visit_date=visit.check_in_timestamp.strftime("%Y-%m-%d"),
                first_name=patient.given_name.get('en'),
                surname=patient.surname.get('en'),
                age=self.age_string_from_dob(patient.date_of_birth),
                date_of_birth=patient.date_of_birth,
                gender=patient.sex,
                home_country=patient.country.get('en'),
                phone=patient.phone
            )
            patient_details_event = patient_details(visit.patient_id)
            if patient_details_event is not None:
                write_patient_details_event(row, patient_details_event)
            for event in events_by_visit(visit.id):
                if event.event_type == 'Medical History':
                    write_medical_hx_event(row, event)
                elif event.event_type == 'Patient Details':
                    write_patient_details_event(row, event)
                elif event.event_type == 'Vitals':
                    write_vitals_event(row, event)
                elif event.event_type == 'Evaluation':
                    write_evaluation_event(row, event)
                elif event.event_type == 'Notes':
                    self.write_text_event(row, 'notes', event.event_metadata)
                elif event.event_type == 'Medicines in Stock':
                    write_med_stock_event(row, event)
                elif event.event_type == 'Medicines OTC':
                    write_med_otc_event(row, event)
                elif event.event_type == 'Controlled Medicines':
                    write_controlled_med_event(row, event)
                elif event.event_type == 'Medical Pathologies':
                    write_med_pathologies_event(row, event)
                elif event.event_type == 'Psychological Pathologies':
                    write_psych_pathologies_event(row, event)
                elif event.event_type == 'Household Environment':
                    write_household_event(row, event)
                elif event.event_type == 'Lab Orders':
                    write_lab_orders_event(row, event)
                elif event.event_type == 'Lab Tests':
                    write_lab_tests_event(row, event)
                elif event.event_type == 'Urine Tests':
                    write_urine_tests_event(row, event)
                elif event.event_type == 'PAP Results':
                    write_pap_event(row, event)
                elif event.event_type == 'Ultrasound':
                    write_ultrasound_event(row, event)
                elif event.event_type == 'Family Planning':
                    write_family_planning_event(row, event)
                elif event.event_type == 'Dental Origin':
                    write_dental_origin_event(row, event)
                elif event.event_type == 'Dental Treatment':
                    write_dental_treatment_event(row, event)
                elif event.event_type == 'Program Trainings':
                    write_program_trainings_event(row, event)
                elif event.event_type == 'Xray Orders':
                    write_xray_orders_event(row, event)
                elif event.event_type == 'Xray Results':
                    write_xray_results_event(row, event)
                elif event.event_type == 'Optometry':
                    write_optometry_event(row, event)
                elif event.event_type == 'Accident Report':
                    write_accident_report_event(row, event)
                elif event.event_type == 'Nursing Care':
                    write_nursing_care_event(row, event)
            yield row

    def write_text_event(self, row, key, text):
        setattr(row, key, text)

    def age_string_from_dob(self, dob):
        if dob is None:
            return 'unknown'
        age = datetime.now() - datetime(dob.year, dob.month, dob.day)
        if age < timedelta(days=365):
            return f'{(age.days // 30) + 1} months'
        return f'{(age.days // 365)} years'
